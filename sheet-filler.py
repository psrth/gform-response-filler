# importing libraries for excel operations
import openpyxl
from openpyxl import Workbook

# importing libraries for randomizing stuff
import random


# setup workbook
wb = Workbook()
ws = wb.active

# define the paramaters for the 2-D array
rows = int(input("Enter number of entries you want: "))
col = int(input("Enter number of datapoints you want: "))

current_column = 0;

# each iteration populates an entire column
while (col>0):
    # column headers
    col_headers = input("Enter header: ")

    # responses to us
    datapoints = input("Add possible entries as one string, seperated by spaces: ")
    datapoints = datapoints.split(' ')
    col = col-1
    current_column = current_column+1

    # populate each row
    for i in range(rows):
        ws.cell(row=i+1, column=current_column, value=random.choice(datapoints))
    
    ws.cell(row=1, column=current_column, value=col_headers)


# save the file locally (FIX THE PATH FOR USE)
wb.save("~/desktop/formresponses.xlsx")